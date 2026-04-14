"""MODE — Biomarker XLSX Parser (openpyxl, no pandas)"""

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


def parse_biomarker_xlsx(filepath: str) -> dict:
    """Parse FOXO biomarker mastersheet into structured data."""
    if not load_workbook:
        return {'sections': {}, 'all_markers': [], 'non_optimal': [], 'status_counts': {}, 'total_markers': 0, 'non_optimal_count': 0}
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {'sections': {}, 'all_markers': [], 'non_optimal': [], 'status_counts': {}, 'total_markers': 0, 'non_optimal_count': 0}

    # Find header row
    headers = [str(c).strip() if c else '' for c in rows[0]]
    col_map = {}
    for i, h in enumerate(headers):
        hl = h.lower()
        if 'biomarker' in hl: col_map['biomarker'] = i
        elif 'value' in hl and 'optimal' not in hl: col_map['value'] = i
        elif 'optimal' in hl or 'range' in hl: col_map['range'] = i
        elif 'severity' in hl: col_map['severity'] = i

    sections = {}
    current_section = 'Uncategorized'
    all_markers = []

    for row in rows[1:]:
        biomarker = str(row[col_map.get('biomarker', 0)] or '').strip()
        value = row[col_map.get('value', 1)]
        severity = str(row[col_map.get('severity', 3)] or '').strip()
        optimal = str(row[col_map.get('range', 2)] or '').strip()

        # Section header: value is empty, severity is empty
        if (value is None or str(value).strip() == '') and severity in ('', 'None', 'nan'):
            if biomarker and biomarker not in ('', 'None', 'nan'):
                current_section = biomarker
                sections[current_section] = []
            continue

        if severity in ('', 'None', 'nan'):
            continue

        marker = {
            'biomarker': biomarker,
            'value': value,
            'optimal_range': optimal if optimal not in ('None', 'nan') else '',
            'severity': severity,
            'section': current_section
        }
        sections.setdefault(current_section, []).append(marker)
        all_markers.append(marker)

    wb.close()

    non_optimal = [m for m in all_markers if m['severity'] != 'Optimal']
    status_counts = {}
    for m in all_markers:
        s = m['severity']
        status_counts[s] = status_counts.get(s, 0) + 1

    return {
        'sections': sections, 'all_markers': all_markers, 'non_optimal': non_optimal,
        'status_counts': status_counts, 'total_markers': len(all_markers),
        'non_optimal_count': len(non_optimal)
    }


def format_sheet2_for_prompt(parsed: dict) -> str:
    lines = ["| Biomarker | Value | Foxo Optimal Range | Foxo Severity |",
             "|-----------|-------|--------------------|---------------|"]
    current = None
    for m in parsed['all_markers']:
        if m['section'] != current:
            current = m['section']
            lines.append(f"\n=== {current} ===")
        lines.append(f"| {m['biomarker']} | {m['value']} | {m['optimal_range']} | {m['severity']} |")
    return '\n'.join(lines)
