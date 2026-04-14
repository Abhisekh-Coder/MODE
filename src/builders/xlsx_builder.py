"""MODE — XLSX Workbook Builder"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

S = {
    'title': Font(name='Arial', bold=True, size=14, color='1F3864'),
    'hdr': Font(name='Arial', bold=True, size=11, color='FFFFFF'),
    'hdr_fill': PatternFill('solid', fgColor='2F5496'),
    'sec': Font(name='Arial', bold=True, size=12, color='1F3864'),
    'sec_fill': PatternFill('solid', fgColor='D6E4F0'),
    'nf': Font(name='Arial', size=10),
    'wrap': Alignment(wrap_text=True, vertical='top'),
    'bdr': Border(left=Side(style='thin', color='B4C6E7'), right=Side(style='thin', color='B4C6E7'),
                  top=Side(style='thin', color='B4C6E7'), bottom=Side(style='thin', color='B4C6E7')),
    'strained': PatternFill('solid', fgColor='C00000'),
    'compensating': PatternFill('solid', fgColor='BF8F00'),
    'stable': PatternFill('solid', fgColor='548235'),
}

def build_workbook(pipeline, output_path: str) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _sheet1(wb, pipeline)
    _sheet2(wb, pipeline)
    if 'agent1' in pipeline.outputs: _sheet3(wb, pipeline.outputs['agent1'])
    if 'agent2' in pipeline.outputs: _sheet4(wb, pipeline.outputs['agent2'])
    if 'agent3' in pipeline.outputs: _sheet5(wb, pipeline.outputs['agent3'])
    wb.save(output_path)
    return output_path

def _sheet1(wb, p):
    ws = wb.create_sheet("Sheet 1 - Input Summary")
    ws['A1'] = f"MODE Playbook — {p.member.get('name','')}"
    ws['A1'].font = S['title']
    r = 3
    for k, v in p.member.items():
        ws.cell(row=r, column=1, value=k.title()); ws.cell(row=r, column=2, value=str(v)); r += 1
    if 'biomarkers' in p.data:
        r += 1; ws.cell(row=r, column=1, value="Total markers"); ws.cell(row=r, column=2, value=p.data['biomarkers']['total_markers'])
        r += 1; ws.cell(row=r, column=1, value="Non-optimal"); ws.cell(row=r, column=2, value=p.data['biomarkers']['non_optimal_count'])
        r += 1
        for status, count in p.data['biomarkers']['status_counts'].items():
            ws.cell(row=r, column=1, value=status); ws.cell(row=r, column=2, value=count); r += 1

def _sheet2(wb, p):
    ws = wb.create_sheet("Sheet 2 - Raw Biomarkers")
    if 'biomarkers' not in p.data: return
    hdrs = ['Biomarker','Value','Optimal Range','Severity']
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h); cell.font = S['hdr']; cell.fill = S['hdr_fill']
    for i, m in enumerate(p.data['biomarkers']['all_markers'], 2):
        ws.cell(row=i, column=1, value=m['biomarker'])
        ws.cell(row=i, column=2, value=m['value'])
        ws.cell(row=i, column=3, value=m['optimal_range'])
        ws.cell(row=i, column=4, value=m['severity'])

def _sheet3(wb, a1):
    ws = wb.create_sheet("Sheet 3 - Analysis")
    for col, w in [('A',42),('B',18),('C',18),('D',16),('E',95)]:
        ws.column_dimensions[col].width = w
    r = 1
    for sec in a1.get('sections', []):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.cell(row=r, column=1, value=sec['name']).font = S['sec']
        ws.cell(row=r, column=1).fill = S['sec_fill']; r += 1
        for c, h in enumerate(['Biomarker','Value','Range','Status','Implication'], 1):
            cell = ws.cell(row=r, column=c, value=h); cell.font = S['hdr']; cell.fill = S['hdr_fill']; r += 1 if c == 5 else 0
        r += 1
        for m in sec.get('markers', []):
            for c, v in enumerate([m['biomarker'], m['value_with_units'], m['optimal_range'], m['status'], m['implication']], 1):
                cell = ws.cell(row=r, column=c, value=v); cell.font = S['nf']; cell.alignment = S['wrap']
            ws.row_dimensions[r].height = 80; r += 1

def _sheet4(wb, a2):
    ws = wb.create_sheet("Sheet 4 - System Mapping")
    for col, w in [('A',6),('B',28),('C',18),('D',80),('E',65),('F',50),('G',55)]:
        ws.column_dimensions[col].width = w
    r = 1
    for c, h in enumerate(['#','System','State','Key Insights','Root Cause','Clinical','Clarity Card'], 1):
        ws.cell(row=r, column=c, value=h).font = S['hdr']; ws.cell(row=r, column=c).fill = S['hdr_fill']
    for sys in a2.get('systems', []):
        r += 1; fill = S.get(sys['state'].lower(), S['strained'])
        ws.cell(row=r, column=1, value=sys['number'])
        ws.cell(row=r, column=2, value=sys['name']); ws.cell(row=r, column=2).fill = fill
        ws.cell(row=r, column=3, value=f"{sys['state']}\n{sys['protocol']}")
        ws.cell(row=r, column=4, value=sys['key_insights'])
        ws.cell(row=r, column=5, value=sys['root_cause'])
        ws.cell(row=r, column=6, value=sys['clinical_implications'])
        ws.cell(row=r, column=7, value=sys['clarity_card'])
        for c in range(1, 8): ws.cell(row=r, column=c).alignment = S['wrap']
        ws.row_dimensions[r].height = 450

def _sheet5(wb, a3):
    ws = wb.create_sheet("Sheet 5 - Humanized Roadmap")
    # Build from Agent 3 structured output — phases and biweekly cards
    ws['A1'] = "Sheet 5 — Humanized Roadmap"
    ws['A1'].font = S['title']
    if 'raw_text' in a3:
        ws['A3'] = a3['raw_text'][:32767]  # Excel cell limit
        ws['A3'].alignment = S['wrap']
